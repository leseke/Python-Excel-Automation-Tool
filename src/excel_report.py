from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

HEADER_FILL = "1F4E78"
KPI_FILL = "D9EAF7"


def _style_header(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(horizontal="center")


def _autofit(ws) -> None:
    for column_cells in ws.columns:
        width = max(len(str(cell.value or "")) for cell in column_cells) + 2
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(width, 35)


def _write_dataframe(ws, frame: pd.DataFrame) -> None:
    ws.append(list(frame.columns))
    for row in frame.itertuples(index=False, name=None):
        ws.append(list(row))
    _style_header(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _autofit(ws)

    if ws.max_row > 1 and ws.max_column > 0:
        table = Table(displayName=f"Table{ws.title.replace(' ', '')}", ref=ws.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)


def build_excel_report(frame: pd.DataFrame, output_path: str | Path) -> Path:
    """Create a client-ready Excel workbook from consolidated sales data."""
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    dashboard = wb.active
    dashboard.title = "Dashboard"

    data_ws = wb.create_sheet("Consolidated Data")
    export = frame.copy()
    if "date" in export.columns:
        export["date"] = export["date"].dt.strftime("%Y-%m-%d")
    _write_dataframe(data_ws, export)

    exceptions = export[export["status"] == "Review"].copy()
    exceptions_ws = wb.create_sheet("Exceptions")
    _write_dataframe(exceptions_ws, exceptions)

    valid = frame[frame["status"] == "Valid"].copy()
    total_records = len(frame)
    valid_records = len(valid)
    review_records = total_records - valid_records
    total_quantity = float(valid["quantity"].sum()) if not valid.empty else 0
    total_revenue = float(valid["revenue"].sum()) if not valid.empty else 0
    unique_products = int(valid["product"].nunique()) if not valid.empty else 0

    dashboard["A1"] = "Automated Sales Report"
    dashboard["A1"].font = Font(size=20, bold=True)
    dashboard.merge_cells("A1:F1")

    kpis = [
        ("Total Records", total_records),
        ("Valid Records", valid_records),
        ("Records to Review", review_records),
        ("Total Quantity", total_quantity),
        ("Total Revenue", total_revenue),
        ("Unique Products", unique_products),
    ]
    for index, (label, value) in enumerate(kpis, start=1):
        cell = dashboard.cell(row=3, column=index, value=label)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor=KPI_FILL)
        cell.alignment = Alignment(horizontal="center")
        value_cell = dashboard.cell(row=4, column=index, value=value)
        value_cell.alignment = Alignment(horizontal="center")
    dashboard["E4"].number_format = '#,##0.00'

    region_summary = (
        valid.groupby("region", dropna=False)["revenue"].sum().sort_values(ascending=False)
        if not valid.empty else pd.Series(dtype=float)
    )
    dashboard["A7"] = "Revenue by Region"
    dashboard["A7"].font = Font(bold=True)
    dashboard.append(["Region", "Revenue"])
    for region, revenue in region_summary.items():
        dashboard.append([region, float(revenue)])
    _style_header(dashboard, 8)

    if len(region_summary):
        bar = BarChart()
        bar.title = "Revenue by Region"
        bar.y_axis.title = "Revenue"
        bar.x_axis.title = "Region"
        bar.add_data(Reference(dashboard, min_col=2, min_row=8, max_row=8 + len(region_summary)), titles_from_data=True)
        bar.set_categories(Reference(dashboard, min_col=1, min_row=9, max_row=8 + len(region_summary)))
        dashboard.add_chart(bar, "D7")

    product_summary = (
        valid.groupby("product")["quantity"].sum().sort_values(ascending=False).head(8)
        if not valid.empty else pd.Series(dtype=float)
    )
    product_start = 10 + len(region_summary)
    dashboard.cell(product_start, 1, "Top Products by Quantity").font = Font(bold=True)
    dashboard.cell(product_start + 1, 1, "Product")
    dashboard.cell(product_start + 1, 2, "Quantity")
    _style_header(dashboard, product_start + 1)
    for offset, (product, quantity) in enumerate(product_summary.items(), start=product_start + 2):
        dashboard.cell(offset, 1, product)
        dashboard.cell(offset, 2, float(quantity))

    if len(product_summary):
        chart = DoughnutChart()
        chart.title = "Top Products"
        chart.add_data(Reference(dashboard, min_col=2, min_row=product_start + 1, max_row=product_start + 1 + len(product_summary)), titles_from_data=True)
        chart.set_categories(Reference(dashboard, min_col=1, min_row=product_start + 2, max_row=product_start + 1 + len(product_summary)))
        dashboard.add_chart(chart, f"D{product_start}")

    monthly = valid.dropna(subset=["date"]).copy()
    if not monthly.empty:
        monthly["month"] = monthly["date"].dt.to_period("M").astype(str)
        monthly_summary = monthly.groupby("month")["revenue"].sum()
        monthly_start = product_start + max(len(product_summary), 8) + 3
        dashboard.cell(monthly_start, 1, "Monthly Revenue").font = Font(bold=True)
        dashboard.cell(monthly_start + 1, 1, "Month")
        dashboard.cell(monthly_start + 1, 2, "Revenue")
        _style_header(dashboard, monthly_start + 1)
        for offset, (month, revenue) in enumerate(monthly_summary.items(), start=monthly_start + 2):
            dashboard.cell(offset, 1, month)
            dashboard.cell(offset, 2, float(revenue))
        line = LineChart()
        line.title = "Monthly Revenue Trend"
        line.add_data(Reference(dashboard, min_col=2, min_row=monthly_start + 1, max_row=monthly_start + 1 + len(monthly_summary)), titles_from_data=True)
        line.set_categories(Reference(dashboard, min_col=1, min_row=monthly_start + 2, max_row=monthly_start + 1 + len(monthly_summary)))
        dashboard.add_chart(line, f"D{monthly_start}")

    dashboard.freeze_panes = "A3"
    for col in range(1, 7):
        dashboard.column_dimensions[get_column_letter(col)].width = 20

    wb.save(output)
    return output
